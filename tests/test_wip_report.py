from datetime import date
from io import BytesIO
import unittest

from openpyxl import load_workbook

from qcc_reflex_pilot.wip_report import (
    MODEL_SHEETS,
    build_wip_rollforward_workbook,
    calculate_rollforward,
)


class WipReportTest(unittest.TestCase):
    def test_rollforward_preserves_floor_and_reports_shortfall_and_excess(self):
        result = calculate_rollforward(
            strains=["Diamond Bar"],
            month_keys=["2026-09", "2026-10"],
            opening_wip={"Diamond Bar": 10.0},
            scheduled={"Diamond Bar": {"2026-10": 60.0}},
            requested={"Diamond Bar": {"2026-09": 12.0, "2026-10": 10.0}},
            minimum_floor_lbs=5.0,
            excess_threshold_lbs=50.0,
        )["Diamond Bar"]

        self.assertEqual(result["released"], [-5.0, -10.0])
        self.assertEqual(result["closing"], [5.0, 55.0])
        self.assertEqual(result["reduction"], [7.0, 0.0])
        self.assertEqual(result["excess"], [0.0, 5.0])

    def test_workbook_contains_four_auditable_model_sheets(self):
        months = [date(2026, 9, 1), date(2026, 10, 1)]
        requested = {
            model: {"Diamond Bar": {"2026-09": 12.0, "2026-10": 10.0}}
            for model in MODEL_SHEETS
        }
        content = build_wip_rollforward_workbook(
            months=months,
            strains=["Diamond Bar"],
            opening_wip={"Diamond Bar": 10.0},
            scheduled={"Diamond Bar": {"2026-10": 60.0}},
            requested_by_model=requested,
            as_of="2026-09-03 12:00 PDT",
        )

        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, list(MODEL_SHEETS.values()))
        sheet = workbook["Current Velocity"]
        self.assertEqual(sheet["A1"].value, "QCC Cultivation WIP Roll-Forward")
        self.assertEqual(sheet["B2"].value, "Current SKU Velocity")
        self.assertEqual(sheet["A10"].value, "Sku")
        self.assertEqual(sheet["B5"].value, 0.0)
        self.assertEqual(sheet["D10"].value, "Unit")
        self.assertEqual(sheet["D11"].value, "lbs")
        self.assertGreaterEqual(sheet.column_dimensions["E"].width, 15)
        self.assertEqual(sheet.freeze_panes, "E11")
        self.assertTrue(any(
            isinstance(cell.value, str) and cell.value.startswith("=MAX(0,")
            for row in sheet.iter_rows()
            for cell in row
        ))
