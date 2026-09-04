"""CFO-ready cultivation WIP roll-forward workbook export."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


MODEL_SHEETS = {
    "Current SKU Velocity": "Current Velocity",
    "30-Day Availability-Adjusted": "30-Day Adjusted",
    "60-Day Availability-Adjusted": "60-Day Adjusted",
    "AI-Adjusted": "AI Adjusted",
}

GREEN = "2F6B4F"
PALE_GREEN = "E2F0D9"
NAVY = "17365D"
PALE_BLUE = "D9EAF7"
PALE_RED = "FCE4D6"
WHITE = "FFFFFF"
GRID = "B7C9C0"
TEXT = "1F2937"


def calculate_rollforward(
    *,
    strains: list[str],
    month_keys: list[str],
    opening_wip: dict[str, float],
    scheduled: dict[str, dict[str, float]],
    requested: dict[str, dict[str, float]],
    minimum_floor_lbs: float = 0.0,
    excess_threshold_lbs: float = 50.0,
) -> dict[str, dict[str, list[float]]]:
    """Calculate the auditable monthly WIP roll-forward for every strain."""
    result: dict[str, dict[str, list[float]]] = {}
    for strain in strains:
        opening_values: list[float] = []
        harvested_values: list[float] = []
        requested_values: list[float] = []
        released_values: list[float] = []
        closing_values: list[float] = []
        reduction_values: list[float] = []
        excess_values: list[float] = []
        opening = max(0.0, float(opening_wip.get(strain, 0.0) or 0.0))
        for month_key in month_keys:
            harvested = max(
                0.0, float(scheduled.get(strain, {}).get(month_key, 0.0) or 0.0)
            )
            request = max(
                0.0, float(requested.get(strain, {}).get(month_key, 0.0) or 0.0)
            )
            releasable = max(0.0, opening + harvested - minimum_floor_lbs)
            release_amount = min(request, releasable)
            released = -release_amount
            closing = max(0.0, opening + harvested + released)
            opening_values.append(round(opening, 2))
            harvested_values.append(round(harvested, 2))
            requested_values.append(round(request, 2))
            released_values.append(round(released, 2))
            closing_values.append(round(closing, 2))
            reduction_values.append(round(max(0.0, request + released), 2))
            excess_values.append(
                round(max(0.0, closing - excess_threshold_lbs), 2)
            )
            opening = closing
        result[strain] = {
            "opening": opening_values,
            "harvested": harvested_values,
            "requested": requested_values,
            "released": released_values,
            "closing": closing_values,
            "reduction": reduction_values,
            "excess": excess_values,
        }
    return result


def _style_header(cell, fill: str = NAVY) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_section_header(
    ws, row: int, title: str, months: list[date], *, main_table: bool = False
) -> None:
    if main_table:
        ws.cell(row=row, column=1, value="Sku")
        ws.cell(row=row, column=2, value="Source")
        _style_header(ws.cell(row=row, column=1))
        _style_header(ws.cell(row=row, column=2))
    else:
        ws.cell(row=row, column=1, value=title)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _style_header(ws.cell(row=row, column=1))
    ws.cell(row=row, column=3, value="Measure")
    _style_header(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="Unit")
    _style_header(ws.cell(row=row, column=4))
    for offset, month in enumerate(months, start=5):
        cell = ws.cell(row=row, column=offset, value=month.strftime("%b %Y"))
        _style_header(cell)


def _write_model_sheet(
    ws,
    *,
    model_name: str,
    months: list[date],
    strains: list[str],
    opening_wip: dict[str, float],
    scheduled: dict[str, dict[str, float]],
    requested: dict[str, dict[str, float]],
    as_of: str,
    minimum_floor_lbs: float,
    excess_threshold_lbs: float,
) -> None:
    month_keys = [month.strftime("%Y-%m") for month in months]
    values = calculate_rollforward(
        strains=strains,
        month_keys=month_keys,
        opening_wip=opening_wip,
        scheduled=scheduled,
        requested=requested,
        minimum_floor_lbs=minimum_floor_lbs,
        excess_threshold_lbs=excess_threshold_lbs,
    )
    last_col = 4 + len(months)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(1, 1, "QCC Cultivation WIP Roll-Forward")
    title.fill = PatternFill("solid", fgColor=GREEN)
    title.font = Font(color=WHITE, bold=True, size=16)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    assumptions = [
        ("Demand model", model_name),
        ("Product scope", "Flower + Pre-Rolls"),
        ("Opening WIP", "Cultivation WIP + Cultivation Pre-WIP"),
        ("Minimum WIP floor", minimum_floor_lbs),
        ("Excess threshold", excess_threshold_lbs),
        ("As of", as_of),
    ]
    for row, (label, value) in enumerate(assumptions, start=2):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color=TEXT)
        ws.cell(row=row, column=2, value=value)
    ws.cell(8, 1, "Scheduled supply uses net dry-flower pounds on the expected availability date (30 days after harvest), after Fresh Frozen and Creative Use reductions. Curing/trim loss is zero because scheduled yield is already net dry flower.")
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=last_col)
    ws.cell(8, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(8, 1).font = Font(italic=True, color="52606D", size=9)
    ws.row_dimensions[8].height = 30

    row = 10
    _write_section_header(ws, row, "WIP Roll-Forward", months, main_table=True)
    rollforward_rows: dict[tuple[str, str], int] = {}
    requested_header_row = 11 + (6 * len(strains))
    actual_header_row = requested_header_row + len(strains) + 1
    reduction_header_row = actual_header_row + len(strains) + 1
    excess_header_row = reduction_header_row + len(strains) + 1
    strain_position = {strain: index for index, strain in enumerate(strains)}
    for strain in strains:
        for measure in (
            "Opening WIP",
            "Harvested in from plants",
            "Curing/trim loss",
            "Released to finished goods",
            "Closing WIP",
        ):
            row += 1
            rollforward_rows[(strain, measure)] = row
            ws.cell(row, 1, strain)
            ws.cell(row, 2, "Cultivation")
            ws.cell(row, 3, measure)
            ws.cell(row, 4, "lbs")
            if measure == "Closing WIP":
                for column in range(1, last_col + 1):
                    ws.cell(row, column).fill = PatternFill("solid", fgColor=PALE_GREEN)
            for index, month_key in enumerate(month_keys, start=5):
                position = index - 5
                if measure == "Opening WIP":
                    if position == 0:
                        value: Any = values[strain]["opening"][position]
                    else:
                        closing_row = rollforward_rows[(strain, "Opening WIP")] + 4
                        value = f"={get_column_letter(index - 1)}{closing_row}"
                elif measure == "Harvested in from plants":
                    value = values[strain]["harvested"][position]
                elif measure == "Curing/trim loss":
                    value = 0
                elif measure == "Released to finished goods":
                    letter = get_column_letter(index)
                    requested_row = requested_header_row + 1 + strain_position[strain]
                    opening_row = rollforward_rows[(strain, "Opening WIP")]
                    harvested_row = rollforward_rows[(strain, "Harvested in from plants")]
                    loss_row = rollforward_rows[(strain, "Curing/trim loss")]
                    value = (
                        f"=-MIN({letter}{requested_row},MAX(0,{letter}{opening_row}"
                        f"+{letter}{harvested_row}-{letter}{loss_row}-$B$5))"
                    )
                else:
                    opening_row = rollforward_rows[(strain, "Opening WIP")]
                    harvested_row = rollforward_rows[(strain, "Harvested in from plants")]
                    loss_row = rollforward_rows[(strain, "Curing/trim loss")]
                    release_row = rollforward_rows[(strain, "Released to finished goods")]
                    letter = get_column_letter(index)
                    value = (
                        f"=MAX(0,{letter}{opening_row}+{letter}{harvested_row}"
                        f"-{letter}{loss_row}+{letter}{release_row})"
                    )
                ws.cell(row, index, value)
        row += 1

    section_specs = [
        ("Requested Release to Finished Goods", "requested", PALE_BLUE),
        ("Actual Release to Finished Goods", "released", PALE_GREEN),
        ("Reduction from Requested Release", "reduction", PALE_RED),
        ("Excess Inventory Above Threshold", "excess", "FFF2CC"),
    ]
    for title_text, value_key, fill in section_specs:
        row += 1
        _write_section_header(ws, row, title_text, months)
        for strain in strains:
            row += 1
            ws.cell(row, 1, strain)
            ws.cell(row, 2, "Cultivation")
            ws.cell(row, 3, "Pounds")
            ws.cell(row, 4, "lbs")
            for column, calculated_value in enumerate(
                values[strain][value_key], start=5
            ):
                letter = get_column_letter(column)
                position = strain_position[strain]
                if value_key == "requested":
                    value: Any = calculated_value
                elif value_key == "released":
                    source_row = rollforward_rows[
                        (strain, "Released to finished goods")
                    ]
                    value = f"={letter}{source_row}"
                elif value_key == "reduction":
                    requested_row = requested_header_row + 1 + position
                    actual_row = actual_header_row + 1 + position
                    value = f"=MAX(0,{letter}{requested_row}+{letter}{actual_row})"
                else:
                    closing_row = rollforward_rows[(strain, "Closing WIP")]
                    value = f"=MAX(0,{letter}{closing_row}-$B$6)"
                ws.cell(row, column, value)
            if value_key in {"reduction", "excess"}:
                for column in range(1, last_col + 1):
                    ws.cell(row, column).fill = PatternFill("solid", fgColor=fill)

    thin = Side(style="thin", color=GRID)
    for data_row in ws.iter_rows(min_row=10, max_row=row, min_col=1, max_col=last_col):
        for cell in data_row:
            cell.border = Border(bottom=thin)
            if cell.column >= 5:
                cell.number_format = '#,##0.0;[Red](#,##0.0);-'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 31
    ws.column_dimensions["D"].width = 9
    for column in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(column)].width = 12
    ws.freeze_panes = "E11"
    ws.auto_filter.ref = f"A10:{get_column_letter(last_col)}{row}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
    ws.oddFooter.center.text = "QCC Control Tower · Cultivation WIP Report"
    ws.oddFooter.right.text = "Page &P of &N"


def build_wip_rollforward_workbook(
    *,
    months: list[date],
    strains: list[str],
    opening_wip: dict[str, float],
    scheduled: dict[str, dict[str, float]],
    requested_by_model: dict[str, dict[str, dict[str, float]]],
    as_of: str | None = None,
    minimum_floor_lbs: float = 0.0,
    excess_threshold_lbs: float = 50.0,
) -> bytes:
    """Build the four-sheet WIP workbook and return its XLSX bytes."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    as_of_label = as_of or datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
    for model_name, sheet_name in MODEL_SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        _write_model_sheet(
            sheet,
            model_name=model_name,
            months=months,
            strains=strains,
            opening_wip=opening_wip,
            scheduled=scheduled,
            requested=requested_by_model.get(model_name, {}),
            as_of=as_of_label,
            minimum_floor_lbs=minimum_floor_lbs,
            excess_threshold_lbs=excess_threshold_lbs,
        )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
